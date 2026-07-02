import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

def create_excel_dashboard(processed_csv_path="data/processed/cleaned_sales.csv", 
                            output_excel_path="dashboard/Amazon Dashboard.xlsx"):
    """
    Builds a professional, stylized Excel Dashboard with KPI cards, summary tables, and charts.
    """
    print(f"Reading processed data from: {processed_csv_path}...")
    df = pd.read_csv(processed_csv_path)
    df['OrderDate'] = pd.to_datetime(df['OrderDate'])
    
    # 1. Initialize Workbook
    wb = openpyxl.Workbook()
    
    # Setup sheets
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_data = wb.create_sheet(title="Cleaned_Data")
    ws_data.views.sheetView[0].showGridLines = True
    
    # 2. Write first 5000 rows of Cleaned Data to Cleaned_Data sheet (for size/performance)
    print("Writing data to Cleaned_Data sheet...")
    # Write headers
    headers = list(df.columns)
    ws_data.append(headers)
    
    # Write rows (limit to 10,000 to keep excel file responsive and compact)
    sample_data = df.head(10000).values.tolist()
    for row in sample_data:
        ws_data.append(row)
        
    # Apply header styling to Cleaned_Data sheet
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws_data.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        
    # Auto-fit columns for Cleaned_Data
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # 3. Create Dashboard Sheet
    print("Creating Dashboard UI...")
    
    # Color Palette Definitions
    FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Deep Blue
    FILL_KPI_BG = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")  # Very Light Grey
    FILL_ACCENT = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Pale Green
    
    FONT_TITLE = Font(name="Segoe UI", size=18, bold=True, color="FFFFFF")
    FONT_SUBTITLE = Font(name="Segoe UI", size=11, italic=True, color="FFFFFF")
    FONT_SECTION = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
    FONT_KPI_VAL = Font(name="Segoe UI", size=16, bold=True, color="2F4F4F")
    FONT_KPI_LBL = Font(name="Segoe UI", size=9, bold=True, color="595959")
    FONT_TABLE_HDR = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    FONT_DATA = Font(name="Segoe UI", size=10, color="000000")
    
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
    ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
    
    BORDER_THIN = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    BORDER_DOUBLE = Border(bottom=Side(style='double', color='1F4E79'), top=Side(style='thin', color='BFBFBF'))
    
    # A. Header Banner (Rows 1-3)
    ws_dash.merge_cells("A1:Q2")
    title_cell = ws_dash["A1"]
    title_cell.value = "AMAZON SALES BUSINESS INTELLIGENCE DASHBOARD"
    title_cell.fill = FILL_HEADER
    title_cell.font = FONT_TITLE
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # B. Set up KPI Calculations from df
    total_sales = df['TotalAmount'].sum()
    total_profit = df['Profit'].sum()
    total_orders = df['OrderID'].nunique()
    avg_sales = df['TotalAmount'].mean()
    profit_margin = total_profit / total_sales
    
    # C. KPI Cards Layout (Row 4 to 6)
    kpis = [
        ("TOTAL SALES", total_sales, "₹#,##0.00", "B", "C"),
        ("TOTAL PROFIT", total_profit, "₹#,##0.00", "E", "F"),
        ("TOTAL ORDERS", total_orders, "#,##0", "H", "I"),
        ("AVG ORDER VALUE", avg_sales, "₹#,##0.00", "K", "L"),
        ("NET MARGIN", profit_margin, "0.00%", "N", "O")
    ]
    
    for label, val, num_fmt, start_col, end_col in kpis:
        # Merge cell ranges
        lbl_range = f"{start_col}4:{end_col}4"
        val_range = f"{start_col}5:{end_col}6"
        
        ws_dash.merge_cells(lbl_range)
        ws_dash.merge_cells(val_range)
        
        lbl_cell = ws_dash[f"{start_col}4"]
        lbl_cell.value = label
        lbl_cell.font = FONT_KPI_LBL
        lbl_cell.fill = FILL_KPI_BG
        lbl_cell.alignment = ALIGN_CENTER
        
        val_cell = ws_dash[f"{start_col}5"]
        val_cell.value = val
        val_cell.font = FONT_KPI_VAL
        val_cell.fill = FILL_KPI_BG
        val_cell.alignment = ALIGN_CENTER
        val_cell.number_format = num_fmt
        
        # Apply borders to the KPI card merged region
        for r in range(4, 7):
            for c in [openpyxl.utils.column_index_from_string(start_col), openpyxl.utils.column_index_from_string(end_col)]:
                ws_dash.cell(row=r, column=c).border = BORDER_THIN
                
    # D. Aggregate and write summary tables (Dashboard Data source)
    print("Writing KPI tables to Dashboard sheet...")
    
    # Table 1: Sales by Category (Row 9-16, Columns B to D)
    ws_dash["B8"] = "Sales by Product Category"
    ws_dash["B8"].font = FONT_SECTION
    
    cat_summary = df.groupby('Category').agg(
        Sales=('TotalAmount', 'sum'),
        Profit=('Profit', 'sum')
    ).reset_index().sort_values(by='Sales', ascending=False)
    
    ws_dash.cell(row=9, column=2, value="Category").font = FONT_TABLE_HDR
    ws_dash.cell(row=9, column=2).fill = FILL_HEADER
    ws_dash.cell(row=9, column=3, value="Sales").font = FONT_TABLE_HDR
    ws_dash.cell(row=9, column=3).fill = FILL_HEADER
    ws_dash.cell(row=9, column=4, value="Profit").font = FONT_TABLE_HDR
    ws_dash.cell(row=9, column=4).fill = FILL_HEADER
    
    row_idx = 10
    for idx, row in cat_summary.iterrows():
        ws_dash.cell(row=row_idx, column=2, value=row['Category']).font = FONT_DATA
        ws_dash.cell(row=row_idx, column=2).border = BORDER_THIN
        
        c_sales = ws_dash.cell(row=row_idx, column=3, value=row['Sales'])
        c_sales.font = FONT_DATA
        c_sales.number_format = "₹#,##0.00"
        c_sales.border = BORDER_THIN
        
        c_profit = ws_dash.cell(row=row_idx, column=4, value=row['Profit'])
        c_profit.font = FONT_DATA
        c_profit.number_format = "₹#,##0.00"
        c_profit.border = BORDER_THIN
        
        row_idx += 1
        
    # Table 2: Sales by State (Row 9-23, Columns F to H)
    ws_dash["F8"] = "Top States by Sales"
    ws_dash["F8"].font = FONT_SECTION
    
    state_summary = df.groupby('State').agg(
        Sales=('TotalAmount', 'sum'),
        Profit=('Profit', 'sum')
    ).reset_index().sort_values(by='Sales', ascending=False).head(10)
    
    ws_dash.cell(row=9, column=6, value="State").font = FONT_TABLE_HDR
    ws_dash.cell(row=9, column=6).fill = FILL_HEADER
    ws_dash.cell(row=9, column=7, value="Sales").font = FONT_TABLE_HDR
    ws_dash.cell(row=9, column=7).fill = FILL_HEADER
    ws_dash.cell(row=9, column=8, value="Profit").font = FONT_TABLE_HDR
    ws_dash.cell(row=9, column=8).fill = FILL_HEADER
    
    row_idx = 10
    for idx, row in state_summary.iterrows():
        ws_dash.cell(row=row_idx, column=6, value=row['State']).font = FONT_DATA
        ws_dash.cell(row=row_idx, column=6).border = BORDER_THIN
        
        c_sales = ws_dash.cell(row=row_idx, column=7, value=row['Sales'])
        c_sales.font = FONT_DATA
        c_sales.number_format = "₹#,##0.00"
        c_sales.border = BORDER_THIN
        
        c_profit = ws_dash.cell(row=row_idx, column=8, value=row['Profit'])
        c_profit.font = FONT_DATA
        c_profit.number_format = "₹#,##0.00"
        c_profit.border = BORDER_THIN
        
        row_idx += 1
        
    # Table 3: Monthly Trend (Row 9-21, Columns J to L)
    ws_dash["J8"] = "Monthly Sales Trend (Last 12 Months)"
    ws_dash["J8"].font = FONT_SECTION
    
    df['MonthYear'] = df['OrderDate'].dt.to_period('M').astype(str)
    monthly_summary = df.groupby('MonthYear').agg(
        Sales=('TotalAmount', 'sum'),
        Profit=('Profit', 'sum')
    ).reset_index().sort_values(by='MonthYear').tail(12)
    
    ws_dash.cell(row=9, column=10, value="Month").font = FONT_TABLE_HDR
    ws_dash.cell(row=9, column=10).fill = FILL_HEADER
    ws_dash.cell(row=9, column=11, value="Sales").font = FONT_TABLE_HDR
    ws_dash.cell(row=9, column=11).fill = FILL_HEADER
    ws_dash.cell(row=9, column=12, value="Profit").font = FONT_TABLE_HDR
    ws_dash.cell(row=9, column=12).fill = FILL_HEADER
    
    row_idx = 10
    for idx, row in monthly_summary.iterrows():
        ws_dash.cell(row=row_idx, column=10, value=row['MonthYear']).font = FONT_DATA
        ws_dash.cell(row=row_idx, column=10).border = BORDER_THIN
        
        c_sales = ws_dash.cell(row=row_idx, column=11, value=row['Sales'])
        c_sales.font = FONT_DATA
        c_sales.number_format = "₹#,##0.00"
        c_sales.border = BORDER_THIN
        
        c_profit = ws_dash.cell(row=row_idx, column=12, value=row['Profit'])
        c_profit.font = FONT_DATA
        c_profit.number_format = "₹#,##0.00"
        c_profit.border = BORDER_THIN
        
        row_idx += 1
        
    # E. Add Excel Charts to Dashboard
    print("Embedding native Excel Charts...")
    
    # 1. Pie Chart for Category Sales
    pie = PieChart()
    labels_ref = Reference(ws_dash, min_col=2, min_row=10, max_row=15)
    data_ref = Reference(ws_dash, min_col=3, min_row=9, max_row=15)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    pie.title = "Sales by Category"
    pie.width = 11
    pie.height = 7
    ws_dash.add_chart(pie, "B18")
    
    # 2. Bar Chart for State Sales
    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "Top 10 States by Sales"
    bar.y_axis.title = "State"
    bar.x_axis.title = "Sales ($)"
    labels_ref_state = Reference(ws_dash, min_col=6, min_row=10, max_row=19)
    data_ref_state = Reference(ws_dash, min_col=7, min_row=9, max_row=19)
    bar.add_data(data_ref_state, titles_from_data=True)
    bar.set_categories(labels_ref_state)
    bar.legend = None
    bar.width = 11
    bar.height = 7
    ws_dash.add_chart(bar, "F21")
    
    # 3. Line Chart for Monthly Sales Trend
    line = LineChart()
    line.title = "Monthly Sales Trend"
    line.style = 13
    line.y_axis.title = "Sales ($)"
    line.x_axis.title = "Month"
    labels_ref_month = Reference(ws_dash, min_col=10, min_row=10, max_row=21)
    data_ref_month = Reference(ws_dash, min_col=11, min_row=9, max_row=21)
    line.add_data(data_ref_month, titles_from_data=True)
    line.set_categories(labels_ref_month)
    line.legend = None
    line.width = 13
    line.height = 7
    ws_dash.add_chart(line, "J23")
    
    # F. Slicers Instruction Panel (Row 25-29, Columns B to H)
    ws_dash.merge_cells("B26:H29")
    slicer_panel = ws_dash["B26"]
    slicer_panel.value = (
        "💡 USER NOTICE (INTERACTIVE SLICERS):\n"
        "To enable interactive slicers (State, Category, Payment Mode) in Microsoft Excel:\n"
        "1. Select the Cleaned_Data tab and press Ctrl+T to convert the data range into an Excel Table.\n"
        "2. Click Table Design > Insert Slicer, and check State, Category, and PaymentMode.\n"
        "3. Copy-paste the slicer boxes into this Dashboard sheet for interactive filtering."
    )
    slicer_panel.font = Font(name="Segoe UI", size=9, italic=True, color="595959")
    slicer_panel.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    slicer_panel.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Pale yellow warning
    
    # Style B26 border
    for r in range(26, 30):
        for c in range(2, 9):
            ws_dash.cell(row=r, column=c).border = BORDER_THIN

    # Set column widths for Dashboard to fit contents nicely
    ws_dash.column_dimensions["A"].width = 3
    ws_dash.column_dimensions["B"].width = 25
    ws_dash.column_dimensions["C"].width = 15
    ws_dash.column_dimensions["D"].width = 15
    ws_dash.column_dimensions["E"].width = 3
    ws_dash.column_dimensions["F"].width = 15
    ws_dash.column_dimensions["G"].width = 15
    ws_dash.column_dimensions["H"].width = 15
    ws_dash.column_dimensions["I"].width = 3
    ws_dash.column_dimensions["J"].width = 15
    ws_dash.column_dimensions["K"].width = 15
    ws_dash.column_dimensions["L"].width = 15
    
    # Save Workbook
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    print(f"Saving Excel Dashboard workbook to: {output_excel_path}...")
    wb.save(output_excel_path)
    print("Excel dashboard created successfully.")

if __name__ == "__main__":
    create_excel_dashboard()
